"""
DebtBusters Intelligence Platform — ML Validation Report
Trains lightweight versions of all 5 ML models on the generated data
Produces: ROC curves, confusion matrices, feature importance, cross-val scores
Output: ../charts/ml_*.png  +  ../excel/ML_Validation_Report.xlsx
Run: python ml_validation_report.py  (after generate_millions.py)
"""

import os, sys, warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.patches import Patch
warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.dirname(__file__))
from brand_colors import CONFLUENT, RAINBOW, DEBTBUSTERS, apply_confluent_style

from sklearn.ensemble import (RandomForestClassifier, GradientBoostingClassifier,
                              GradientBoostingRegressor)
from sklearn.multioutput import MultiOutputRegressor
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.metrics import (roc_auc_score, roc_curve, classification_report,
                              confusion_matrix, mean_absolute_error, r2_score,
                              average_precision_score, precision_recall_curve)
from sklearn.calibration import CalibratedClassifierCV
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

apply_confluent_style()

DATA_DIR  = os.path.join(os.path.dirname(__file__), "..", "data")
CHART_DIR = os.path.join(os.path.dirname(__file__), "..", "charts")
EXCEL_DIR = os.path.join(os.path.dirname(__file__), "..", "excel")

def load(name):
    pp = os.path.join(DATA_DIR, f"{name}.parquet")
    pc = os.path.join(DATA_DIR, f"{name}.csv")
    return pd.read_parquet(pp) if os.path.exists(pp) else pd.read_csv(pc)

def savefig(name):
    path = os.path.join(CHART_DIR, f"{name}.png")
    plt.savefig(path, bbox_inches="tight", facecolor=DEBTBUSTERS["light_grey"], dpi=150)
    print(f"  Saved: {name}.png")
    plt.close("all")

def argb(h): return "FF" + h.lstrip("#").upper()
def fill(h): return PatternFill("solid", fgColor=argb(h))
def hdr(bold=True, color="FFFFFFFF"):
    return Font(bold=bold, size=10, color=color, name="Calibri")
def align(h="center"):
    return Alignment(horizontal=h, vertical="center")

NAVY   = DEBTBUSTERS["navy"]
TEAL   = CONFLUENT["teal"]
GREEN  = CONFLUENT["green"]
RED    = CONFLUENT["red"]
BLUE   = CONFLUENT["blue"]
PURPLE = CONFLUENT["purple"]
ORANGE = CONFLUENT["orange"]

# ── load & sample for speed (full dataset would be used in Databricks) ─────────
print("Loading data (sampling for local ML validation)…")
SAMPLE = 50_000   # sample size per model — representative

leads    = load("leads")
assess   = load("assessments")
cases    = load("debt_review_cases")
pays     = load("payments")
credit   = load("credit_monitoring")
clients  = load("clients")
plans    = load("repayment_plans")
# risk_score is a causal signal injected during data generation
if "risk_score" not in clients.columns:
    clients["risk_score"] = 0.3   # fallback if old data

def enc(df, cols):
    for c in cols:
        if c in df.columns:
            le = LabelEncoder()
            df[c+"_enc"] = le.fit_transform(df[c].fillna("UNKNOWN").astype(str))
    return df

results = {}   # will hold per-model metrics

# ── MODEL 1: Lead Conversion ───────────────────────────────────────────────────
print("\n[1/5] Lead Conversion Prediction…")
df1 = leads.merge(
    assess[["client_id","gross_income","debt_to_income_ratio","over_indebted_flag","number_of_creditors"]].drop_duplicates("client_id"),
    on="client_id", how="left"
).merge(
    clients[["client_id","age","gender","employment_status","income_band","province","risk_score"]].drop_duplicates("client_id"),
    on="client_id", how="left"
)
df1 = enc(df1, ["source_channel","utm_source","employment_status","income_band","province","gender"])
feat1 = ["lead_score","cost_per_lead","gross_income","debt_to_income_ratio",
         "number_of_creditors","age","over_indebted_flag","risk_score",
         "source_channel_enc","utm_source_enc","employment_status_enc",
         "income_band_enc","province_enc","gender_enc"]
feat1 = [f for f in feat1 if f in df1.columns]
df1 = df1[feat1+["converted_flag"]].dropna().sample(min(SAMPLE,len(df1)), random_state=42)
X1,y1 = df1[feat1].fillna(0), df1["converted_flag"].astype(int)
X1_tr,X1_te,y1_tr,y1_te = train_test_split(X1,y1,test_size=0.2,random_state=42,stratify=y1)

m1 = RandomForestClassifier(n_estimators=100, max_depth=8, class_weight="balanced",
                             random_state=42, n_jobs=-1)
m1.fit(X1_tr, y1_tr)
y1_proba = m1.predict_proba(X1_te)[:,1]
y1_pred  = (y1_proba>=0.5).astype(int)
auc1     = roc_auc_score(y1_te, y1_proba)
aps1     = average_precision_score(y1_te, y1_proba)
fpr1,tpr1,_ = roc_curve(y1_te, y1_proba)
cv1 = cross_val_score(m1, X1.fillna(0), y1, cv=5, scoring="roc_auc")
results["Lead Conversion"] = {"auc":auc1,"aps":aps1,"cv_mean":cv1.mean(),"cv_std":cv1.std(),
                               "fpr":fpr1,"tpr":tpr1,"y_te":y1_te,"y_pred":y1_pred,
                               "features":feat1,"importance":m1.feature_importances_,
                               "cm":confusion_matrix(y1_te,y1_pred)}
print(f"   AUC={auc1:.4f} | 5-Fold CV={cv1.mean():.4f} ± {cv1.std():.4f}")

# ── MODEL 2: Payment Default ───────────────────────────────────────────────────
print("[2/5] Payment Default Prediction…")
pay_agg = (pays.groupby("client_id").agg(
    total_pays=("payment_id","count"),
    missed=("missed_payment_flag","sum"),
    avg_coll=("collection_rate","mean"),
    total_arr=("arrears_amount","sum"),
).reset_index())
pay_agg["miss_rate"] = pay_agg["missed"] / pay_agg["total_pays"]

df2 = pays.sample(min(SAMPLE,len(pays)),random_state=42).merge(
    pay_agg, on="client_id", how="left"
).merge(
    clients[["client_id","age","employment_status","income_band","province","risk_score"]].drop_duplicates("client_id"),
    on="client_id", how="left"
)
df2 = enc(df2,["employment_status","income_band","province"])
feat2 = ["total_pays","missed","avg_coll","total_arr","miss_rate","age","payment_year",
         "risk_score","employment_status_enc","income_band_enc","province_enc"]
feat2 = [f for f in feat2 if f in df2.columns]
df2 = df2[feat2+["missed_payment_flag"]].dropna()
X2,y2 = df2[feat2].fillna(0), df2["missed_payment_flag"].astype(int)
X2_tr,X2_te,y2_tr,y2_te = train_test_split(X2,y2,test_size=0.2,random_state=42,stratify=y2)

m2 = GradientBoostingClassifier(n_estimators=100, max_depth=5, learning_rate=0.1, random_state=42)
m2.fit(X2_tr, y2_tr)
y2_proba = m2.predict_proba(X2_te)[:,1]
y2_pred  = (y2_proba>=0.4).astype(int)
auc2     = roc_auc_score(y2_te, y2_proba)
aps2     = average_precision_score(y2_te, y2_proba)
fpr2,tpr2,_ = roc_curve(y2_te, y2_proba)
cv2 = cross_val_score(m2, X2.fillna(0), y2, cv=5, scoring="roc_auc")
results["Payment Default"] = {"auc":auc2,"aps":aps2,"cv_mean":cv2.mean(),"cv_std":cv2.std(),
                               "fpr":fpr2,"tpr":tpr2,"y_te":y2_te,"y_pred":y2_pred,
                               "features":feat2,"importance":m2.feature_importances_,
                               "cm":confusion_matrix(y2_te,y2_pred)}
print(f"   AUC={auc2:.4f} | 5-Fold CV={cv2.mean():.4f} ± {cv2.std():.4f}")

# ── MODEL 3: Product Recommendation ──────────────────────────────────────────
print("[3/5] Product Recommendation…")
df3 = assess.sample(min(SAMPLE,len(assess)),random_state=42).merge(
    clients[["client_id","age","gender","employment_status","income_band","province"]].drop_duplicates("client_id"),
    on="client_id", how="left"
)
df3 = enc(df3,["gender","employment_status","income_band","province","dti_band"])
le3 = LabelEncoder()
df3["target"] = le3.fit_transform(df3["recommended_product"].fillna("Credit Monitoring"))
feat3 = ["gross_income","net_income","living_expenses","total_debt_balance",
         "total_monthly_debt_instalment","disposable_income","debt_to_income_ratio",
         "affordability_amount","number_of_creditors","age",
         "gender_enc","employment_status_enc","income_band_enc","province_enc","dti_band_enc"]
feat3 = [f for f in feat3 if f in df3.columns]
df3 = df3[feat3+["target"]].dropna()
X3,y3 = df3[feat3].fillna(0), df3["target"]
X3_tr,X3_te,y3_tr,y3_te = train_test_split(X3,y3,test_size=0.2,random_state=42,stratify=y3)

m3 = RandomForestClassifier(n_estimators=150, max_depth=10, class_weight="balanced",
                             random_state=42, n_jobs=-1)
m3.fit(X3_tr, y3_tr)
y3_pred = m3.predict(X3_te)
acc3    = (y3_pred==y3_te).mean()
cv3     = cross_val_score(m3, X3.fillna(0), y3, cv=5, scoring="accuracy")
results["Product Recommendation"] = {"acc":acc3,"cv_mean":cv3.mean(),"cv_std":cv3.std(),
                                      "features":feat3,"importance":m3.feature_importances_,
                                      "classes":le3.classes_,
                                      "cm":confusion_matrix(y3_te,y3_pred)}
print(f"   Accuracy={acc3:.4f} | 5-Fold CV={cv3.mean():.4f} ± {cv3.std():.4f}")

# ── MODEL 4: Credit Score Forecast ────────────────────────────────────────────
print("[4/5] Credit Score Improvement Forecast…")
cr = credit.copy()
cr["monitoring_date"] = pd.to_datetime(cr["monitoring_date"])
cr = cr.sort_values(["client_id","monitoring_date"])
# Predict future scores using rolling mean + seq-based trajectory
cr["roll3"]     = cr.groupby("client_id")["credit_score"].transform(lambda x:x.rolling(3,min_periods=1).mean())
# Targets: expected score after 3/6/12 months of improvement (seq * 3 pts/month)
cr["score_3m"]  = (cr["credit_score"] + cr["monitoring_seq"] * 3 * 3 +
                   cr.groupby("client_id")["score_change"].transform("mean") * 3).clip(300, 850)
cr["score_6m"]  = (cr["credit_score"] + cr["monitoring_seq"] * 3 * 6 +
                   cr.groupby("client_id")["score_change"].transform("mean") * 6).clip(300, 850)
cr["score_12m"] = (cr["credit_score"] + cr["monitoring_seq"] * 3 * 12 +
                   cr.groupby("client_id")["score_change"].transform("mean") * 12).clip(300, 850)
cr = cr.merge(clients[["client_id","age","employment_status","income_band"]].drop_duplicates("client_id"),
               on="client_id", how="left")
cr = enc(cr,["credit_risk_band","bureau","employment_status","income_band"])
feat4 = ["credit_score","score_change","monitoring_seq","accounts_in_arrears","judgements_count",
         "defaults_count","enquiries_count","total_utilisation_pct","roll3","age",
         "credit_risk_band_enc","bureau_enc","employment_status_enc","income_band_enc"]
feat4 = [f for f in feat4 if f in cr.columns]
tgt4  = ["score_3m","score_6m","score_12m"]
cr4   = cr.dropna(subset=tgt4).sample(min(SAMPLE,len(cr.dropna(subset=tgt4))),random_state=42)
X4,y4 = cr4[feat4].fillna(0), cr4[tgt4]
X4_tr,X4_te,y4_tr,y4_te = train_test_split(X4,y4,test_size=0.2,random_state=42)

m4 = MultiOutputRegressor(
    GradientBoostingRegressor(n_estimators=100,learning_rate=0.1,max_depth=5,random_state=42),
    n_jobs=-1
)
m4.fit(X4_tr, y4_tr)
y4_pred = m4.predict(X4_te)
mae4  = [mean_absolute_error(y4_te.iloc[:,i], y4_pred[:,i]) for i in range(3)]
r2_4  = [r2_score(y4_te.iloc[:,i], y4_pred[:,i]) for i in range(3)]
fi4   = np.mean([e.feature_importances_ for e in m4.estimators_], axis=0)
results["Credit Score Forecast"] = {"mae":mae4,"r2":r2_4,"features":feat4,"importance":fi4,
                                     "y_te":y4_te.values,"y_pred":y4_pred,"targets":tgt4}
for t,m,r in zip(tgt4,mae4,r2_4):
    print(f"   {t}: MAE={m:.1f} | R²={r:.4f}")

# ── MODEL 5: Client Churn ─────────────────────────────────────────────────────
print("[5/5] Client Churn Prediction…")
df5 = cases.merge(
    clients[["client_id","age","employment_status","income_band","province","risk_score"]].drop_duplicates("client_id"),
    on="client_id", how="left"
).merge(
    assess[["client_id","debt_to_income_ratio","total_debt_balance","disposable_income",
            "over_indebted_flag","number_of_creditors"]].drop_duplicates("client_id"),
    on="client_id", how="left"
)
df5["target"] = (df5["case_stage"]=="Withdrawn").astype(int)
df5 = enc(df5,["employment_status","income_band","province","legal_status","ncr_status"])
feat5 = ["age","debt_to_income_ratio","total_debt_balance","disposable_income",
         "number_of_creditors","days_in_stage","risk_score",
         "employment_status_enc","income_band_enc","province_enc","legal_status_enc","ncr_status_enc"]
feat5 = [f for f in feat5 if f in df5.columns]
df5 = df5[feat5+["target"]].dropna()
X5,y5 = df5[feat5].fillna(0), df5["target"]
X5_tr,X5_te,y5_tr,y5_te = train_test_split(X5,y5,test_size=0.2,random_state=42,stratify=y5)

m5 = GradientBoostingClassifier(n_estimators=150,max_depth=6,learning_rate=0.08,random_state=42)
m5.fit(X5_tr, y5_tr)
y5_proba = m5.predict_proba(X5_te)[:,1]
y5_pred  = (y5_proba>=0.4).astype(int)
auc5     = roc_auc_score(y5_te, y5_proba)
aps5     = average_precision_score(y5_te, y5_proba)
fpr5,tpr5,_ = roc_curve(y5_te, y5_proba)
cv5 = cross_val_score(m5, X5.fillna(0), y5, cv=5, scoring="roc_auc")
results["Client Churn"] = {"auc":auc5,"aps":aps5,"cv_mean":cv5.mean(),"cv_std":cv5.std(),
                            "fpr":fpr5,"tpr":tpr5,"y_te":y5_te,"y_pred":y5_pred,
                            "features":feat5,"importance":m5.feature_importances_,
                            "cm":confusion_matrix(y5_te,y5_pred)}
print(f"   AUC={auc5:.4f} | 5-Fold CV={cv5.mean():.4f} ± {cv5.std():.4f}")

# ══════════════════════════════════════════════════════════════════════════════
# CHART ML-01: Combined ROC Curves (4 classifiers)
# ══════════════════════════════════════════════════════════════════════════════
print("\nGenerating ML charts…")

fig, axes = plt.subplots(2, 2, figsize=(13, 10))
fig.patch.set_facecolor(DEBTBUSTERS["light_grey"])
fig.suptitle("DebtBusters ML Models — ROC Curves", fontsize=16,
             fontweight="bold", color=DEBTBUSTERS["navy"], y=1.01)

binary_models = ["Lead Conversion","Payment Default","Client Churn"]
bi_colors     = [CONFLUENT["blue"], CONFLUENT["red"], CONFLUENT["purple"]]
bi_axes       = [axes[0,0], axes[0,1], axes[1,0]]

for ax,(name,color) in zip(bi_axes, zip(binary_models,bi_colors)):
    r = results[name]
    ax.set_facecolor("#FFFFFF")
    ax.plot(r["fpr"], r["tpr"], color=color, linewidth=2.5,
            label=f"AUC = {r['auc']:.4f}")
    ax.fill_between(r["fpr"], r["tpr"], alpha=0.12, color=color)
    ax.plot([0,1],[0,1],"--",color=DEBTBUSTERS["mid_grey"],linewidth=1,label="Random")
    ax.set_xlabel("False Positive Rate"); ax.set_ylabel("True Positive Rate")
    ax.set_title(f"{name}", fontsize=12, fontweight="bold", color=DEBTBUSTERS["navy"])
    ax.legend(fontsize=10); ax.set_xlim(0,1); ax.set_ylim(0,1.02)
    ax.spines[["top","right"]].set_visible(False)
    ax.text(0.60, 0.15, f"CV: {r['cv_mean']:.4f} ± {r['cv_std']:.4f}",
            transform=ax.transAxes, fontsize=9, color=DEBTBUSTERS["navy"])

# Credit Score Forecast subplot
ax4 = axes[1,1]; ax4.set_facecolor("#FFFFFF")
r4  = results["Credit Score Forecast"]
for i,(tgt,color) in enumerate(zip(r4["targets"], [CONFLUENT["blue"],CONFLUENT["teal"],CONFLUENT["purple"]])):
    sample_idx = np.random.choice(len(r4["y_te"]), 500, replace=False)
    ax4.scatter(r4["y_te"][sample_idx,i], r4["y_pred"][sample_idx,i],
                alpha=0.3, s=8, color=color, label=f"{tgt} (R²={r4['r2'][i]:.3f})")
ax4.plot([300,850],[300,850],"--",color=DEBTBUSTERS["mid_grey"],linewidth=1)
ax4.set_xlabel("Actual Credit Score"); ax4.set_ylabel("Predicted Credit Score")
ax4.set_title("Credit Score Forecast (3/6/12M)", fontsize=12,
              fontweight="bold", color=DEBTBUSTERS["navy"])
ax4.legend(fontsize=9); ax4.spines[["top","right"]].set_visible(False)

plt.tight_layout()
savefig("ml_01_roc_curves")

# ══════════════════════════════════════════════════════════════════════════════
# CHART ML-02: Feature Importance (top 10 per binary model)
# ══════════════════════════════════════════════════════════════════════════════
fig, axes = plt.subplots(1, 3, figsize=(17, 6))
fig.patch.set_facecolor(DEBTBUSTERS["light_grey"])
fig.suptitle("Feature Importance — Top 10 Features per Model",
             fontsize=14, fontweight="bold", color=DEBTBUSTERS["navy"])

for ax,(name,color) in zip(axes, zip(binary_models, bi_colors)):
    r   = results[name]
    fi  = pd.Series(r["importance"], index=r["features"]).nlargest(10)
    ax.set_facecolor("#FFFFFF")
    bars = ax.barh(fi.index[::-1], fi.values[::-1], color=color, alpha=0.85,
                   edgecolor="white", height=0.6)
    for bar in bars:
        ax.text(bar.get_width()+0.001, bar.get_y()+bar.get_height()/2,
                f"{bar.get_width():.3f}", va="center", fontsize=8)
    ax.set_title(f"{name}", fontsize=11, fontweight="bold", color=DEBTBUSTERS["navy"])
    ax.set_xlabel("Importance")
    ax.spines[["top","right"]].set_visible(False)

plt.tight_layout()
savefig("ml_02_feature_importance")

# ══════════════════════════════════════════════════════════════════════════════
# CHART ML-03: Confusion Matrices
# ══════════════════════════════════════════════════════════════════════════════
fig, axes = plt.subplots(1, 3, figsize=(15, 4))
fig.patch.set_facecolor(DEBTBUSTERS["light_grey"])
fig.suptitle("Confusion Matrices — Binary Classifiers",
             fontsize=14, fontweight="bold", color=DEBTBUSTERS["navy"])

import matplotlib.colors as mcolors

for ax,(name,color) in zip(axes, zip(binary_models, bi_colors)):
    cm_val = results[name]["cm"]
    cmap = mcolors.LinearSegmentedColormap.from_list("",["#FFFFFF", color])
    im = ax.imshow(cm_val, cmap=cmap, interpolation="nearest")
    ax.set_facecolor("#FFFFFF")
    for i in range(2):
        for j in range(2):
            ax.text(j, i, f"{cm_val[i,j]:,}",
                    ha="center", va="center",
                    color="white" if cm_val[i,j] > cm_val.max()*0.5 else DEBTBUSTERS["navy"],
                    fontsize=14, fontweight="bold")
    ax.set_xticks([0,1]); ax.set_yticks([0,1])
    ax.set_xticklabels(["Predicted 0","Predicted 1"], fontsize=9)
    ax.set_yticklabels(["Actual 0","Actual 1"], fontsize=9)
    ax.set_title(f"{name}", fontsize=11, fontweight="bold", color=DEBTBUSTERS["navy"])

plt.tight_layout()
savefig("ml_03_confusion_matrices")

# ══════════════════════════════════════════════════════════════════════════════
# CHART ML-04: Cross-Validation Comparison
# ══════════════════════════════════════════════════════════════════════════════
fig, ax = plt.subplots(figsize=(10, 5))
fig.patch.set_facecolor(DEBTBUSTERS["light_grey"])
ax.set_facecolor("#FFFFFF")

model_names = ["Lead\nConversion","Payment\nDefault","Product\nRecommendation","Client\nChurn"]
cv_means = [results["Lead Conversion"]["cv_mean"], results["Payment Default"]["cv_mean"],
            results["Product Recommendation"]["cv_mean"], results["Client Churn"]["cv_mean"]]
cv_stds  = [results["Lead Conversion"]["cv_std"],  results["Payment Default"]["cv_std"],
            results["Product Recommendation"]["cv_std"],  results["Client Churn"]["cv_std"]]
colors   = [CONFLUENT["blue"], CONFLUENT["red"], CONFLUENT["green"], CONFLUENT["purple"]]

bars = ax.bar(model_names, cv_means, yerr=cv_stds, color=colors, alpha=0.85,
              edgecolor="white", capsize=6, error_kw={"linewidth":1.5,"ecolor":DEBTBUSTERS["navy"]})
ax.axhline(0.80, color=CONFLUENT["teal"], linestyle="--", linewidth=1.2, label="Good performance (0.80)")
ax.axhline(0.90, color=CONFLUENT["green"],linestyle="--", linewidth=1.2, label="Excellent (0.90)")
for bar, val in zip(bars, cv_means):
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.008,
            f"{val:.4f}", ha="center", va="bottom", fontsize=11, fontweight="bold",
            color=DEBTBUSTERS["navy"])

ax.set_ylim(0.5, 1.05)
ax.set_ylabel("5-Fold Cross-Validation Score (AUC / Accuracy)")
ax.set_title("Model Performance — 5-Fold Cross-Validation Comparison",
             fontsize=14, fontweight="bold", color=DEBTBUSTERS["navy"])
ax.legend(fontsize=9)
ax.spines[["top","right"]].set_visible(False)
savefig("ml_04_cross_validation")

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL Validation Report
# ══════════════════════════════════════════════════════════════════════════════
print("\nWriting ML Validation Excel report…")
wb = Workbook()
ws = wb.active; ws.title = "ML Model Summary"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:H1")
ws["A1"] = "DebtBusters Intelligence Platform — ML Model Validation Report"
ws["A1"].font = Font(bold=True, size=16, color="FFFFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=argb(DEBTBUSTERS["navy"]))
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

ws["A2"] = f"Generated: {pd.Timestamp.now().strftime('%d %B %Y %H:%M')}"
ws["A2"].font = Font(size=9, color=argb(DEBTBUSTERS["mid_grey"]))
ws.row_dimensions[2].height = 18

headers = ["Model","Algorithm","Metric","Score","5-Fold CV Mean","5-Fold CV Std",
           "Avg Precision","Status"]
for j,h in enumerate(headers):
    cell = ws.cell(4, j+1, h)
    cell.font = Font(bold=True, size=10, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=argb(DEBTBUSTERS["navy"]))
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[4].height = 22

model_summary = [
    ("Lead Conversion",        "Random Forest",    "AUC-ROC",
     f"{results['Lead Conversion']['auc']:.4f}",
     f"{results['Lead Conversion']['cv_mean']:.4f}",
     f"{results['Lead Conversion']['cv_std']:.4f}",
     f"{results['Lead Conversion']['aps']:.4f}"),
    ("Payment Default",        "Gradient Boosting","AUC-ROC",
     f"{results['Payment Default']['auc']:.4f}",
     f"{results['Payment Default']['cv_mean']:.4f}",
     f"{results['Payment Default']['cv_std']:.4f}",
     f"{results['Payment Default']['aps']:.4f}"),
    ("Product Recommendation", "Random Forest",    "Accuracy",
     f"{results['Product Recommendation']['acc']:.4f}",
     f"{results['Product Recommendation']['cv_mean']:.4f}",
     f"{results['Product Recommendation']['cv_std']:.4f}",
     "N/A"),
    ("Credit Score Forecast",  "Multi-output GBM", "R² (12M)",
     f"{results['Credit Score Forecast']['r2'][2]:.4f}",
     "N/A","N/A","N/A"),
    ("Client Churn",           "Gradient Boosting","AUC-ROC",
     f"{results['Client Churn']['auc']:.4f}",
     f"{results['Client Churn']['cv_mean']:.4f}",
     f"{results['Client Churn']['cv_std']:.4f}",
     f"{results['Client Churn']['aps']:.4f}"),
]
row_colors = [argb(CONFLUENT["blue"]), argb(CONFLUENT["red"]),
              argb(CONFLUENT["green"]), argb(CONFLUENT["teal"]), argb(CONFLUENT["purple"])]

for row_i,(name,algo,metric,score,cv_m,cv_s,aps_v) in enumerate(model_summary):
    r     = row_i + 5
    score_f = float(score) if score != "N/A" else 0
    status  = "Excellent" if score_f >= 0.85 else "Good" if score_f >= 0.75 else "Acceptable"
    st_col  = argb(GREEN) if status=="Excellent" else argb(ORANGE) if status=="Good" else argb(RED)

    vals = [name,algo,metric,score,cv_m,cv_s,aps_v,status]
    for j,val in enumerate(vals):
        cell = ws.cell(r, j+1, val)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        bg = argb("#F5F6FA") if row_i%2==0 else "FFFFFFFF"
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(size=10)
        if j==0:
            cell.fill = PatternFill("solid", fgColor=row_colors[row_i])
            cell.font = Font(bold=True, size=10, color="FFFFFFFF")
        if j==7:
            cell.fill = PatternFill("solid", fgColor=st_col)
            cell.font = Font(bold=True, size=10, color="FFFFFFFF")
    ws.row_dimensions[r].height = 22

for col,w in zip("ABCDEFGH",[30,24,18,14,16,14,16,14]):
    ws.column_dimensions[col].width = w

# Detail tabs per model
for name, r_dict, color in [
    ("Lead Conversion",    results["Lead Conversion"],    BLUE),
    ("Payment Default",    results["Payment Default"],    RED),
    ("Client Churn",       results["Client Churn"],       PURPLE),
]:
    ws2 = wb.create_sheet(f"Detail — {name[:15]}")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:D1")
    ws2["A1"] = f"Feature Importance — {name}"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor=argb(DEBTBUSTERS["navy"]))
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 28

    fi = pd.Series(r_dict["importance"], index=r_dict["features"]).sort_values(ascending=False)
    for j,h in enumerate(["Feature","Importance","Rank","Cumulative"]):
        c = ws2.cell(3, j+1, h)
        c.font = Font(bold=True, color="FFFFFFFF"); c.fill = PatternFill("solid", fgColor=argb(color))
        c.alignment = Alignment(horizontal="center", vertical="center")

    cumsum = 0
    for row_i,(feat,imp) in enumerate(fi.items()):
        r = row_i+4; cumsum += imp
        bg = argb("#F5F6FA") if row_i%2==0 else "FFFFFFFF"
        for j,val in enumerate([feat, round(imp,5), row_i+1, round(cumsum,4)]):
            cell = ws2.cell(r,j+1,val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center" if j>0 else "left", vertical="center")
            cell.font = Font(size=9)
            if j==0: cell.alignment = Alignment(horizontal="left", vertical="center")

    for col,w in zip("ABCD",[35,15,10,15]):
        ws2.column_dimensions[col].width = w

out_path = os.path.join(EXCEL_DIR, "ML_Validation_Report.xlsx")
wb.save(out_path)
print(f"ML Validation Excel saved: {out_path}")
print(f"\n{'='*60}")
print("  ML VALIDATION COMPLETE")
print(f"  Lead Conversion   AUC: {results['Lead Conversion']['auc']:.4f}")
print(f"  Payment Default   AUC: {results['Payment Default']['auc']:.4f}")
print(f"  Product Reco   Acc:    {results['Product Recommendation']['acc']:.4f}")
print(f"  Credit Forecast R²:    {results['Credit Score Forecast']['r2'][2]:.4f}")
print(f"  Client Churn      AUC: {results['Client Churn']['auc']:.4f}")
print(f"{'='*60}")
