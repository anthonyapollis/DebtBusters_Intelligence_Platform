"""
DebtBusters Intelligence Platform — Excel Report Generator
Creates a fully formatted multi-sheet Excel workbook with charts + KPIs
Run: python generate_excel_report.py  (after generate_millions.py)
Output: ../excel/DebtBusters_Intelligence_Report.xlsx
Requirements: pip install openpyxl xlsxwriter
"""

import os, sys, warnings
import numpy as np
import pandas as pd
from datetime import datetime
warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.dirname(__file__))
from brand_colors import CONFLUENT, RAINBOW, DEBTBUSTERS

DATA_DIR  = os.path.join(os.path.dirname(__file__), "..", "data")
EXCEL_DIR = os.path.join(os.path.dirname(__file__), "..", "excel")
os.makedirs(EXCEL_DIR, exist_ok=True)
OUT_PATH  = os.path.join(EXCEL_DIR, "DebtBusters_Intelligence_Report.xlsx")

# ── helpers ───────────────────────────────────────────────────────────────────
def load(name):
    pp = os.path.join(DATA_DIR, f"{name}.parquet")
    pc = os.path.join(DATA_DIR, f"{name}.csv")
    return pd.read_parquet(pp) if os.path.exists(pp) else pd.read_csv(pc)

def hex2rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2],16) for i in (0,2,4))

def col_letter(n):           # 1-based column index → Excel letter
    s = ""
    while n:
        n, r = divmod(n-1, 26)
        s = chr(65+r) + s
    return s

# openpyxl colors are ARGB strings (no #)
def argb(hex_color):
    return "FF" + hex_color.lstrip("#").upper()

# ── load data ─────────────────────────────────────────────────────────────────
print("Loading data…")
leads    = load("leads")
assess   = load("assessments")
cases    = load("debt_review_cases")
pays     = load("payments")
credit   = load("credit_monitoring")
clients  = load("clients")
creditors= load("creditors")
plans    = load("repayment_plans")

# ── computed summaries ────────────────────────────────────────────────────────
leads["lead_date"] = pd.to_datetime(leads["lead_date"])
pays["payment_date"] = pd.to_datetime(pays["payment_date"])

kpis = {
    "Total Clients":           len(clients),
    "Total Leads":             len(leads),
    "Qualified Leads":         int(leads["qualified_flag"].sum()),
    "Assessed Leads":          int(leads["assessed_flag"].sum()),
    "Converted Leads":         int(leads["converted_flag"].sum()),
    "Lead Conversion Rate":    f"{leads['converted_flag'].mean()*100:.2f}%",
    "Total Assessments":       len(assess),
    "Over-Indebted Clients":   int(assess["over_indebted_flag"].sum()),
    "Over-Indebted Rate":      f"{assess['over_indebted_flag'].mean()*100:.2f}%",
    "Avg DTI Ratio":           f"{assess['debt_to_income_ratio'].mean()*100:.2f}%",
    "Avg Total Debt":          f"R{assess['total_debt_balance'].mean():,.0f}",
    "Total Cases":             len(cases),
    "Active Cases":            int((cases["case_stage"]=="Active").sum()),
    "Completed Cases":         int((cases["case_stage"]=="Completed").sum()),
    "Withdrawn Cases":         int((cases["case_stage"]=="Withdrawn").sum()),
    "Clearances Issued":       int(cases["clearance_issued_flag"].sum()),
    "Case Completion Rate":    f"{(cases['case_stage']=='Completed').mean()*100:.2f}%",
    "Avg Days in Stage":       f"{cases['days_in_stage'].mean():.1f}",
    "Total Payments":          len(pays),
    "Collection Rate":         f"{pays['actual_payment_amount'].sum()/pays['expected_payment_amount'].sum()*100:.2f}%",
    "Missed Payment Rate":     f"{pays['missed_payment_flag'].mean()*100:.2f}%",
    "Total Expected Payments": f"R{pays['expected_payment_amount'].sum()/1e6:.1f}M",
    "Total Actual Payments":   f"R{pays['actual_payment_amount'].sum()/1e6:.1f}M",
    "Total Arrears":           f"R{pays['arrears_amount'].sum()/1e6:.2f}M",
    "Avg Credit Score":        f"{credit['credit_score'].mean():.0f}",
    "Creditor Acceptance Rate":f"{plans['accepted_flag'].mean()*100:.2f}%",
    "Avg Monthly Saving":      f"R{plans[plans['accepted_flag']==True]['monthly_saving'].mean():,.0f}",
    "Total Estimated Savings": f"R{plans[plans['accepted_flag']==True]['total_saving_estimated'].sum()/1e9:.2f}B",
}

# ── write workbook ────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill, numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage

wb = Workbook()

# ── colour helpers ─────────────────────────────────────────────────────────────
NAVY   = argb(DEBTBUSTERS["navy"])
TEAL   = argb(CONFLUENT["teal"])
WHITE  = "FFFFFFFF"
L_GREY = argb(DEBTBUSTERS["light_grey"])
M_GREY = argb(DEBTBUSTERS["mid_grey"])
RED    = argb(CONFLUENT["red"])
GREEN  = argb(CONFLUENT["green"])
ORANGE = argb(CONFLUENT["orange"])
YELLOW = argb(CONFLUENT["yellow"])
BLUE   = argb(CONFLUENT["blue"])
PURPLE = argb(CONFLUENT["purple"])

def fill(hex_arg):     return PatternFill("solid", fgColor=hex_arg)
def font(bold=False, size=10, color="FF000000", name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)
def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color=M_GREY)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bottom():
    t = Side(style="medium", color=NAVY)
    return Border(bottom=t)

# ── SHEET 1: Cover ────────────────────────────────────────────────────────────
ws_cover = wb.active
ws_cover.title = "Cover"
ws_cover.sheet_view.showGridLines = False

ws_cover.column_dimensions["A"].width = 2
for col in ["B","C","D","E","F","G","H"]:
    ws_cover.column_dimensions[col].width = 18

# Title block
ws_cover.merge_cells("B2:H2")
ws_cover["B2"] = "DebtBusters Intelligence Platform"
ws_cover["B2"].font = Font(bold=True, size=28, color=WHITE, name="Calibri")
ws_cover["B2"].fill = fill(NAVY)
ws_cover["B2"].alignment = align("center","center")
ws_cover.row_dimensions[2].height = 50

ws_cover.merge_cells("B3:H3")
ws_cover["B3"] = "Confluent Financial Wellness Group — Enterprise Analytics Report"
ws_cover["B3"].font = Font(bold=False, size=13, color=TEAL, name="Calibri")
ws_cover["B3"].fill = fill(NAVY)
ws_cover["B3"].alignment = align("center","center")
ws_cover.row_dimensions[3].height = 28

# Rainbow stripe
for i, c in enumerate(RAINBOW):
    col = get_column_letter(i+2)
    ws_cover[f"{col}4"] = ""
    ws_cover[f"{col}4"].fill = fill(argb(c))
ws_cover.row_dimensions[4].height = 10

# Tagline
ws_cover.merge_cells("B6:H6")
ws_cover["B6"] = '"Building Financially Healthy Societies, Together"'
ws_cover["B6"].font = Font(italic=True, size=14, color=argb(DEBTBUSTERS["navy"]))
ws_cover["B6"].alignment = align("center","center")

# KPI summary tiles
kpi_tiles = [
    ("Total Clients",       f"{len(clients):,}",        BLUE),
    ("Total Leads",         f"{len(leads):,}",           TEAL),
    ("Conversion Rate",     f"{leads['converted_flag'].mean()*100:.1f}%", GREEN),
    ("Active Cases",        f"{int((cases['case_stage']=='Active').sum()):,}", PURPLE),
    ("Collection Rate",     f"{pays['actual_payment_amount'].sum()/pays['expected_payment_amount'].sum()*100:.1f}%", TEAL),
    ("Total Debt Managed",  f"R{assess['total_debt_balance'].sum()/1e9:.1f}B", ORANGE),
    ("Avg Credit Score",    f"{int(credit['credit_score'].mean())}",            YELLOW),
    ("Est. Total Savings",  f"R{plans[plans['accepted_flag']==True]['total_saving_estimated'].sum()/1e9:.2f}B", GREEN),
]

row_start = 9
for i,(label, val, color) in enumerate(kpi_tiles):
    col = get_column_letter(i%4 * 2 + 2) if i<4 else get_column_letter((i-4)%4 * 2 + 2)
    row = row_start if i<4 else row_start+4

    ws_cover.merge_cells(f"{col}{row}:{get_column_letter(ord(col[0])-64+1)}{row}")
    ws_cover[f"{col}{row}"] = val
    ws_cover[f"{col}{row}"].font = Font(bold=True, size=20, color=WHITE)
    ws_cover[f"{col}{row}"].fill = fill(color)
    ws_cover[f"{col}{row}"].alignment = align("center","center")
    ws_cover.row_dimensions[row].height = 36

    ws_cover.merge_cells(f"{col}{row+1}:{get_column_letter(ord(col[0])-64+1)}{row+1}")
    ws_cover[f"{col}{row+1}"] = label
    ws_cover[f"{col}{row+1}"].font = Font(size=9, color=argb(DEBTBUSTERS["navy"]))
    ws_cover[f"{col}{row+1}"].fill = fill(L_GREY)
    ws_cover[f"{col}{row+1}"].alignment = align("center","center")

# Report date
ws_cover.merge_cells("B20:H20")
ws_cover["B20"] = f"Report Generated: {datetime.now().strftime('%d %B %Y')}  |  Anthony Apollis  |  Confluent DebtBusters DW"
ws_cover["B20"].font = Font(size=9, color=M_GREY)
ws_cover["B20"].alignment = align("center","center")

# ── SHEET 2: Executive KPIs ───────────────────────────────────────────────────
ws_kpi = wb.create_sheet("Executive KPIs")
ws_kpi.sheet_view.showGridLines = False
ws_kpi.column_dimensions["A"].width = 3
ws_kpi.column_dimensions["B"].width = 38
ws_kpi.column_dimensions["C"].width = 25
ws_kpi.column_dimensions["D"].width = 18

ws_kpi.merge_cells("B1:D1")
ws_kpi["B1"] = "Executive KPI Summary"
ws_kpi["B1"].font = Font(bold=True, size=16, color=WHITE)
ws_kpi["B1"].fill = fill(NAVY)
ws_kpi["B1"].alignment = align("left","center")
ws_kpi.row_dimensions[1].height = 32

headers = ["Metric","Value","Category"]
header_colors = [NAVY, NAVY, NAVY]
for i,(h,c) in enumerate(zip(headers, header_colors)):
    cell = ws_kpi.cell(3, i+2, h)
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = fill(c)
    cell.alignment = align("center","center")
ws_kpi.row_dimensions[3].height = 22

category_map = {
    "Total Clients":"Client Base","Total Leads":"Marketing","Qualified Leads":"Marketing",
    "Assessed Leads":"Marketing","Converted Leads":"Marketing","Lead Conversion Rate":"Marketing",
    "Total Assessments":"Operations","Over-Indebted Clients":"Operations",
    "Over-Indebted Rate":"Operations","Avg DTI Ratio":"Operations","Avg Total Debt":"Operations",
    "Total Cases":"Cases","Active Cases":"Cases","Completed Cases":"Cases","Withdrawn Cases":"Cases",
    "Clearances Issued":"Cases","Case Completion Rate":"Cases","Avg Days in Stage":"Cases",
    "Total Payments":"Payments","Collection Rate":"Payments","Missed Payment Rate":"Payments",
    "Total Expected Payments":"Payments","Total Actual Payments":"Payments","Total Arrears":"Payments",
    "Avg Credit Score":"Credit","Creditor Acceptance Rate":"Creditors",
    "Avg Monthly Saving":"Savings","Total Estimated Savings":"Savings",
}
cat_colors = {
    "Client Base": BLUE, "Marketing": TEAL, "Operations": GREEN,
    "Cases": PURPLE, "Payments": ORANGE, "Credit": YELLOW, "Creditors": RED, "Savings": GREEN,
}

zebra = [L_GREY, WHITE]
for row_i,(metric, value) in enumerate(kpis.items()):
    r   = row_i + 4
    cat = category_map.get(metric,"General")
    bg  = zebra[row_i % 2]
    for c, val, align_h in [(2,metric,"left"),(3,value,"center"),(4,cat,"center")]:
        cell = ws_kpi.cell(r, c, val)
        cell.font = font(size=10)
        cell.fill = fill(bg)
        cell.alignment = align(align_h,"center")
        cell.border = thin_border()
    ws_kpi.cell(r,4).fill = fill(cat_colors.get(cat, L_GREY))
    ws_kpi.cell(r,4).font = Font(bold=True, size=9, color=WHITE)
    ws_kpi.row_dimensions[r].height = 18

# ── SHEET 3: Lead Funnel Data ─────────────────────────────────────────────────
ws_lead = wb.create_sheet("Lead Funnel")
ws_lead.sheet_view.showGridLines = False

ws_lead.merge_cells("A1:G1")
ws_lead["A1"] = "Lead Funnel Analysis — Monthly"
ws_lead["A1"].font = Font(bold=True, size=14, color=WHITE)
ws_lead["A1"].fill = fill(NAVY)
ws_lead["A1"].alignment = align("left","center")
ws_lead.row_dimensions[1].height = 28

leads["month_year"] = leads["lead_date"].dt.to_period("M").astype(str)
monthly_leads = leads.groupby("month_year").agg(
    total_leads=("lead_id","count"),
    qualified=("qualified_flag","sum"),
    assessed=("assessed_flag","sum"),
    converted=("converted_flag","sum"),
    avg_cost=("cost_per_lead","mean"),
    avg_score=("lead_score","mean"),
).reset_index()
monthly_leads["conversion_rate"] = (monthly_leads["converted"] / monthly_leads["total_leads"] * 100).round(2)

col_headers = ["Month","Total Leads","Qualified","Assessed","Converted","Conv Rate %","Avg Cost/Lead","Avg Score"]
for j, h in enumerate(col_headers):
    cell = ws_lead.cell(3, j+1, h)
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = fill(NAVY)
    cell.alignment = align("center","center")

monthly_leads.columns = ["month_year","total_leads","qualified","assessed","converted","avg_cost","avg_score","conversion_rate"]
for row_i, row in monthly_leads.iterrows():
    r = row_i + 4
    vals = [row["month_year"], int(row["total_leads"]), int(row["qualified"]),
            int(row["assessed"]), int(row["converted"]),
            round(row["conversion_rate"],2), round(row["avg_cost"],2), round(row["avg_score"],1)]
    for j, val in enumerate(vals):
        cell = ws_lead.cell(r, j+1, val)
        cell.fill = fill(L_GREY if row_i%2==0 else WHITE)
        cell.alignment = align("center","center")
        cell.border = thin_border()

for col in "ABCDEFGH":
    ws_lead.column_dimensions[col].width = 16

# Bar chart — Total Leads per Month
chart_lead = BarChart()
chart_lead.title = "Monthly Lead Volume"
chart_lead.style = 10
chart_lead.height = 12; chart_lead.width = 22
chart_lead.y_axis.title = "Leads"
data_ref  = Reference(ws_lead, min_col=2, max_col=5,
                       min_row=3, max_row=3+len(monthly_leads))
cats_ref  = Reference(ws_lead, min_col=1, min_row=4, max_row=3+len(monthly_leads))
chart_lead.add_data(data_ref, titles_from_data=True)
chart_lead.set_categories(cats_ref)
chart_lead.shape = 4
ws_lead.add_chart(chart_lead, "J3")

# ── SHEET 4: Payment Performance ─────────────────────────────────────────────
ws_pay = wb.create_sheet("Payment Performance")
ws_pay.sheet_view.showGridLines = False

ws_pay.merge_cells("A1:I1")
ws_pay["A1"] = "Monthly Payment Performance — PDA Collections"
ws_pay["A1"].font = Font(bold=True, size=14, color=WHITE)
ws_pay["A1"].fill = fill(NAVY)
ws_pay["A1"].alignment = align("left","center")

pays_m = pays.groupby("payment_month").agg(
    expected=("expected_payment_amount","sum"),
    actual=("actual_payment_amount","sum"),
    arrears=("arrears_amount","sum"),
    missed=("missed_payment_flag","sum"),
    total=("payment_id","count"),
).reset_index()
pays_m["collection_rate"]  = (pays_m["actual"]/pays_m["expected"]*100).round(2)
pays_m["missed_rate"]      = (pays_m["missed"]/pays_m["total"]*100).round(2)
pays_m = pays_m.sort_values("payment_month")

pay_headers = ["Month","Expected (R)","Actual (R)","Arrears (R)",
               "Missed","Total","Collection %","Missed %"]
for j, h in enumerate(pay_headers):
    cell = ws_pay.cell(3, j+1, h)
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = fill(NAVY)
    cell.alignment = align("center","center")

for row_i, row in pays_m.iterrows():
    r = row_i + 4
    vals = [row["payment_month"],
            round(row["expected"],0), round(row["actual"],0), round(row["arrears"],0),
            int(row["missed"]), int(row["total"]),
            round(row["collection_rate"],2), round(row["missed_rate"],2)]
    for j, val in enumerate(vals):
        cell = ws_pay.cell(r, j+1, val)
        cell.fill = fill(L_GREY if row_i%2==0 else WHITE)
        cell.alignment = align("center","center")
        cell.border = thin_border()
        if j == 6:   # collection rate — colour code
            if isinstance(val, float) and val >= 95:
                cell.fill = fill(argb(CONFLUENT["green"]))
            elif isinstance(val, float) and val < 85:
                cell.fill = fill(RED)
                cell.font = Font(bold=True, color=WHITE)

for col in "ABCDEFGHI":
    ws_pay.column_dimensions[col].width = 17

# Line chart
lc = LineChart()
lc.title = "Collection Rate Trend"
lc.style = 10; lc.height = 12; lc.width = 22
lc.y_axis.title = "Rate (%)"; lc.y_axis.numFmt = "0.0"
cr_ref = Reference(ws_pay, min_col=7, min_row=3, max_row=3+len(pays_m))
mr_ref = Reference(ws_pay, min_col=8, min_row=3, max_row=3+len(pays_m))
cat_ref= Reference(ws_pay, min_col=1, min_row=4, max_row=3+len(pays_m))
lc.add_data(cr_ref, titles_from_data=True)
lc.add_data(mr_ref, titles_from_data=True)
lc.set_categories(cat_ref)
ws_pay.add_chart(lc, "J3")

# ── SHEET 5: Case Pipeline ────────────────────────────────────────────────────
ws_case = wb.create_sheet("Case Pipeline")
ws_case.sheet_view.showGridLines = False

ws_case.merge_cells("A1:F1")
ws_case["A1"] = "Debt Review Case Pipeline"
ws_case["A1"].font = Font(bold=True, size=14, color=WHITE)
ws_case["A1"].fill = fill(NAVY)
ws_case["A1"].alignment = align("left","center")

stage_order = ["Lead","Assessment","Application","Review","Court Order","Active","Completed","Withdrawn"]
stage_data  = cases.groupby("case_stage").agg(
    count=("case_id","count"),
    avg_days=("days_in_stage","mean"),
    clearances=("clearance_issued_flag","sum"),
).reindex(stage_order, fill_value=0).reset_index()

case_headers = ["Stage","Count","% of Total","Avg Days in Stage","Clearances Issued"]
for j, h in enumerate(case_headers):
    cell = ws_case.cell(3, j+1, h)
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = fill(NAVY)
    cell.alignment = align("center","center")

total_cases = stage_data["count"].sum()
stage_colors_xl = [BLUE,TEAL,GREEN,YELLOW,ORANGE,PURPLE,TEAL,RED]
for row_i, row in stage_data.iterrows():
    r = row_i + 4
    pct = row["count"]/total_cases*100 if total_cases > 0 else 0
    vals = [row["case_stage"], int(row["count"]), f"{pct:.1f}%",
            f"{row['avg_days']:.1f}", int(row["clearances"])]
    for j, val in enumerate(vals):
        cell = ws_case.cell(r, j+1, val)
        cell.alignment = align("center","center")
        cell.border = thin_border()
        if j == 0:
            cell.fill = fill(stage_colors_xl[row_i % len(stage_colors_xl)])
            cell.font = Font(bold=True, color=WHITE)
        else:
            cell.fill = fill(L_GREY if row_i%2==0 else WHITE)
    ws_case.row_dimensions[r].height = 20

for col in "ABCDEF":
    ws_case.column_dimensions[col].width = 22

# Pie chart for case stages
pc = PieChart()
pc.title = "Case Stage Distribution"; pc.style = 10
pc.height = 14; pc.width = 18
data_ref = Reference(ws_case, min_col=2, min_row=3, max_row=3+len(stage_data))
cats_ref = Reference(ws_case, min_col=1, min_row=4, max_row=3+len(stage_data))
pc.add_data(data_ref, titles_from_data=True)
pc.set_categories(cats_ref)
ws_case.add_chart(pc, "H3")

# ── SHEET 6: Credit Risk ──────────────────────────────────────────────────────
ws_cr = wb.create_sheet("Credit Risk")
ws_cr.sheet_view.showGridLines = False
ws_cr.merge_cells("A1:F1")
ws_cr["A1"] = "Credit Risk & Monitoring Dashboard"
ws_cr["A1"].font = Font(bold=True, size=14, color=WHITE)
ws_cr["A1"].fill = fill(NAVY)
ws_cr["A1"].alignment = align("left","center")

risk_data = credit.groupby("credit_risk_band").agg(
    count=("monitoring_id","count"),
    avg_score=("credit_score","mean"),
    avg_change=("score_change","mean"),
    avg_arrears=("accounts_in_arrears","mean"),
    avg_judgements=("judgements_count","mean"),
).reset_index()
risk_order = ["Very High Risk","High Risk","Medium Risk","Low Risk","Very Low Risk"]
risk_data = risk_data.set_index("credit_risk_band").reindex(risk_order).reset_index()
risk_data.columns = ["Risk Band","Count","Avg Score","Avg Change","Avg Arrears","Avg Judgements"]

cr_headers = list(risk_data.columns)
for j, h in enumerate(cr_headers):
    cell = ws_cr.cell(3, j+1, h)
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = fill(NAVY)
    cell.alignment = align("center","center")

risk_colors_xl = [RED, ORANGE, YELLOW, GREEN, TEAL]
for row_i, row in risk_data.iterrows():
    r = row_i + 4
    vals = [row["Risk Band"], int(row["Count"]) if not pd.isna(row["Count"]) else 0,
            round(row["Avg Score"],1) if not pd.isna(row["Avg Score"]) else 0,
            round(row["Avg Change"],2) if not pd.isna(row["Avg Change"]) else 0,
            round(row["Avg Arrears"],2) if not pd.isna(row["Avg Arrears"]) else 0,
            round(row["Avg Judgements"],2) if not pd.isna(row["Avg Judgements"]) else 0]
    for j, val in enumerate(vals):
        cell = ws_cr.cell(r, j+1, val)
        cell.alignment = align("center","center")
        cell.border = thin_border()
        if j == 0:
            cell.fill = fill(risk_colors_xl[row_i])
            cell.font = Font(bold=True, color=WHITE)
        else:
            cell.fill = fill(L_GREY if row_i%2==0 else WHITE)
    ws_cr.row_dimensions[r].height = 20

for col in "ABCDEFG":
    ws_cr.column_dimensions[col].width = 22

# ── SHEET 7: Raw Sample Data ──────────────────────────────────────────────────
for sheet_name, df, max_rows in [
    ("Sample — Leads",       leads.head(2000),    2000),
    ("Sample — Assessments", assess.head(2000),   2000),
    ("Sample — Payments",    pays.head(2000),      2000),
]:
    ws_raw = wb.create_sheet(sheet_name)
    ws_raw.sheet_view.showGridLines = True
    ws_raw.freeze_panes = "A2"

    for j, col in enumerate(df.columns):
        cell = ws_raw.cell(1, j+1, col)
        cell.font = Font(bold=True, size=10, color=WHITE)
        cell.fill = fill(NAVY)
        cell.alignment = align("center","center")

    for row_i, row in df.iterrows():
        r = row_i + 2
        for j, val in enumerate(row):
            cell = ws_raw.cell(r, j+1)
            if isinstance(val, (np.bool_,)):
                cell.value = bool(val)
            elif isinstance(val, (np.integer,)):
                cell.value = int(val)
            elif isinstance(val, (np.floating,)):
                cell.value = float(val) if not np.isnan(val) else None
            else:
                cell.value = str(val) if val is not None else None
            cell.fill = fill(L_GREY if row_i%2==0 else WHITE)

    for j in range(len(df.columns)):
        ws_raw.column_dimensions[get_column_letter(j+1)].width = 18

print(f"Saving Excel workbook…")
wb.save(OUT_PATH)
print(f"Excel report saved: {OUT_PATH}")
